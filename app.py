"""
app.py — Keeler Prospect Agent
Pantoscope Sales Command Center micro-app.

Auth:   X-SCC-User header (set by shell, not handled here)
Port:   PORT environment variable (required — shell-assigned)
Data:   ICD-10 and HCPCS loaded at startup; CPT via AMA Zip API if key available
LLM:    Claude (Anthropic) with web_search tool
"""

import os
import io
import json
import time
import zipfile
import requests as req_lib
from flask import Flask, request, jsonify, send_from_directory, Response, stream_with_context
from flask_cors import CORS
from dotenv import load_dotenv
import anthropic
import openpyxl

load_dotenv()

app = Flask(__name__, static_folder="static", static_url_path="")
CORS(app)

# ── Anthropic client ───────────────────────────────────────────────────────────
_anthropic_key = os.environ.get("ANTHROPIC_API_KEY")
if not _anthropic_key:
    print("WARNING: ANTHROPIC_API_KEY not set — analysis will fail.")
client = anthropic.Anthropic(api_key=_anthropic_key)
MODEL = os.environ.get("CLAUDE_MODEL", "claude-opus-4-8")


# ══════════════════════════════════════════════════════════════════════════════
# Reference Data — loaded once at startup
# ══════════════════════════════════════════════════════════════════════════════

ICD10: dict = {}
HCPCS_DATA: dict = {}
CPT_DATA: dict = {}

# ── ICD-10-CM ─────────────────────────────────────────────────────────────────
def _load_icd10():
    global ICD10
    path = os.path.join(os.path.dirname(__file__), "data", "icd10cm-codes-2026.txt")
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                line = line.rstrip("\r\n")
                if not line.strip():
                    continue
                parts = line.split(None, 1)
                if len(parts) == 2:
                    ICD10[parts[0].strip()] = parts[1].strip()
        print(f"[startup] ICD-10 loaded: {len(ICD10):,} codes")
    except FileNotFoundError:
        print("[startup] WARNING: ICD-10 data file not found — continuing without it")


def get_icd10_ophthalmic_context() -> str:
    if not ICD10:
        return ""
    ophthalmic = {k: v for k, v in ICD10.items() if k.startswith("H")}
    groups = {
        "Conjunctiva (H10-H13)": [],
        "Cornea (H16-H22)": [],
        "Lens/Cataract (H25-H28)": [],
        "Retina (H30-H36)": [],
        "Glaucoma (H40-H42)": [],
        "Vitreous/Globe (H43-H44)": [],
        "Visual field/acuity (H53-H54)": [],
    }
    for code, desc in ophthalmic.items():
        n = int(code[1:3]) if len(code) >= 3 and code[1:3].isdigit() else -1
        if 10 <= n <= 13:   groups["Conjunctiva (H10-H13)"].append((code, desc))
        elif 16 <= n <= 22: groups["Cornea (H16-H22)"].append((code, desc))
        elif 25 <= n <= 28: groups["Lens/Cataract (H25-H28)"].append((code, desc))
        elif 30 <= n <= 36: groups["Retina (H30-H36)"].append((code, desc))
        elif 40 <= n <= 42: groups["Glaucoma (H40-H42)"].append((code, desc))
        elif 43 <= n <= 44: groups["Vitreous/Globe (H43-H44)"].append((code, desc))
        elif 53 <= n <= 54: groups["Visual field/acuity (H53-H54)"].append((code, desc))

    lines = ["KEY OPHTHALMIC ICD-10-CM CODES (FY2026):"]
    for grp, items in groups.items():
        lines.append(f"\n{grp}:")
        for code, desc in items[:12]:
            display = f"{code[:3]}.{code[3:]}" if len(code) > 3 else code
            lines.append(f"  {display}  {desc}")
        if len(items) > 12:
            lines.append(f"  ... and {len(items)-12} more codes in this category")
    return "\n".join(lines)


# ── HCPCS ─────────────────────────────────────────────────────────────────────
def _load_hcpcs():
    global HCPCS_DATA
    folder = os.path.join(os.path.dirname(__file__), "hcpc2026_jul_anweb_06172026")
    path = os.path.join(folder, "HCPC2026_JUL_ANWEB_06172026.xlsx")
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers = None
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(h).strip() if h else "" for h in row]
                continue
            row_dict = dict(zip(headers, row))
            code = str(row_dict.get("HCPC", "") or "").strip()
            if not code or len(code) < 4:
                continue
            term = str(row_dict.get("TERM DT", "") or "").strip()
            if term and term not in ("", "0", "00000000"):
                continue
            HCPCS_DATA[code] = {
                "short": str(row_dict.get("SHORT DESCRIPTION", "") or "").strip(),
                "long":  str(row_dict.get("LONG DESCRIPTION",  "") or "").strip(),
                "cov":   str(row_dict.get("COV",               "") or "").strip(),
                "betos": str(row_dict.get("BETOS",             "") or "").strip(),
            }
        wb.close()
        print(f"[startup] HCPCS loaded: {len(HCPCS_DATA):,} codes")
    except FileNotFoundError:
        print("[startup] WARNING: HCPCS xlsx not found — continuing without it")
    except Exception as e:
        print(f"[startup] WARNING: HCPCS load failed: {e}")


def get_ophthalmic_hcpcs_context() -> str:
    if not HCPCS_DATA:
        return ""
    OPHTHALMIC_PREFIXES = ("J", "V2", "Q09", "A6")
    lines = ["KEY OPHTHALMIC HCPCS CODES (CMS July 2026):"]
    for code, data in HCPCS_DATA.items():
        if any(code.startswith(p) for p in OPHTHALMIC_PREFIXES):
            cov = " [COVERED]" if data["cov"] == "C" else ""
            short = data["short"] or data["long"]
            if short:
                lines.append(f"  {code}  {short[:80]}{cov}")
    if len(lines) == 1:
        lines.append("  (No relevant codes found in loaded file)")
    return "\n".join(lines[:80])


# ── CPT ───────────────────────────────────────────────────────────────────────
def _load_cpt():
    global CPT_DATA
    cpt_key = os.environ.get("CPT_API_KEY")
    cpt_dir = os.path.join(os.path.dirname(__file__), "data", "cpt-standard")

    if os.path.isdir(cpt_dir) and any(os.listdir(cpt_dir)):
        _parse_cpt_dir(cpt_dir)
        return

    if not cpt_key:
        print("[startup] CPT_API_KEY not set — CPT data unavailable")
        return

    print("[startup] Downloading CPT Standard Data File from AMA API...")
    try:
        r = req_lib.get(
            "https://api-platform.ama-assn.org/cpt-zip/1.0.0",
            headers={"Authorization": f"Bearer {cpt_key}"},
            timeout=60,
        )
        if r.status_code == 401:
            print("[startup] CPT download failed: 401 Unauthorized — token may be expired.")
            print("[startup]   → Update CPT_API_KEY in .env with a fresh token and restart.")
            return
        if r.status_code == 403:
            print("[startup] CPT download failed: 403 Forbidden — check AMA license status.")
            return
        if r.status_code != 200:
            print(f"[startup] CPT download failed: HTTP {r.status_code} — {r.text[:200]}")
            return

        os.makedirs(cpt_dir, exist_ok=True)
        z = zipfile.ZipFile(io.BytesIO(r.content))
        z.extractall(cpt_dir)
        print(f"[startup] CPT zip extracted to {cpt_dir}")
        _parse_cpt_dir(cpt_dir)
    except Exception as e:
        print(f"[startup] CPT load error: {e}")


def _parse_cpt_dir(cpt_dir: str):
    global CPT_DATA
    for fname in os.listdir(cpt_dir):
        fpath = os.path.join(cpt_dir, fname)
        if fname.endswith(".txt") and "long" in fname.lower():
            try:
                with open(fpath, "r", encoding="utf-8", errors="ignore") as f:
                    for line in f:
                        parts = line.split(None, 1)
                        if len(parts) == 2 and parts[0].isdigit() and len(parts[0]) == 5:
                            CPT_DATA[parts[0]] = parts[1].strip()
            except Exception:
                pass
        elif fname.endswith(".xlsx"):
            try:
                wb = openpyxl.load_workbook(fpath, read_only=True)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    if row and row[0] and str(row[0]).isdigit() and len(str(row[0])) == 5:
                        CPT_DATA[str(row[0])] = str(row[1]).strip() if row[1] else ""
                wb.close()
            except Exception:
                pass
    if CPT_DATA:
        print(f"[startup] CPT loaded: {len(CPT_DATA):,} codes")
    else:
        print("[startup] CPT: no parseable file found in extracted zip")


_load_icd10()
_load_hcpcs()
_load_cpt()


# ══════════════════════════════════════════════════════════════════════════════
# NPI Registry Lookup
# ══════════════════════════════════════════════════════════════════════════════

NPI_BASE = "https://npiregistry.cms.hhs.gov/api/?version=2.1"


def lookup_npi(npi_number: str = None, org_name: str = None, state: str = None) -> dict:
    params = {"limit": 5}
    if npi_number:
        params["number"] = npi_number.strip()
    elif org_name:
        params["organization_name"] = org_name.strip()
        if state:
            params["state"] = state.strip().upper()[:2]
    else:
        return {}

    try:
        r = req_lib.get(NPI_BASE, params=params, timeout=6)
        if r.status_code != 200:
            return {}
        data = r.json()
        if not data.get("result_count"):
            return {}

        results = []
        for rec in data["results"]:
            basic = rec.get("basic", {})
            loc_addr = next(
                (a for a in rec.get("addresses", []) if a.get("address_purpose") == "LOCATION"),
                rec.get("addresses", [{}])[0] if rec.get("addresses") else {}
            )
            dba = next(
                (n.get("organization_name", "") for n in rec.get("other_names", [])
                 if n.get("type") == "Doing Business As"),
                ""
            )
            primary_tax = next(
                (t for t in rec.get("taxonomies", []) if t.get("primary")),
                (rec.get("taxonomies") or [{}])[0]
            )
            all_taxonomies = [t.get("desc", "") for t in rec.get("taxonomies", []) if t.get("desc")]
            results.append({
                "npi": rec.get("number"),
                "type": rec.get("enumeration_type"),
                "status": basic.get("status"),
                "org_name": basic.get("organization_name") or
                            f"{basic.get('first_name','')} {basic.get('last_name','')}".strip(),
                "dba": dba,
                "address": f"{loc_addr.get('address_1','')} {loc_addr.get('address_2','')}".strip(),
                "city": loc_addr.get("city", ""),
                "state": loc_addr.get("state", ""),
                "zip": loc_addr.get("postal_code", ""),
                "phone": loc_addr.get("telephone_number", ""),
                "primary_specialty": primary_tax.get("desc", ""),
                "all_specialties": all_taxonomies,
                "enumeration_date": basic.get("enumeration_date", ""),
                "last_updated": basic.get("last_updated", ""),
            })
        return {"found": True, "count": len(results), "records": results}
    except Exception as e:
        print(f"[npi] lookup error: {e}")
        return {}


# ══════════════════════════════════════════════════════════════════════════════
# System Prompt
# ══════════════════════════════════════════════════════════════════════════════

SYSTEM_PROMPT = """You are the Keeler Prospect Intelligence Agent — an ophthalmic sales analyst for Keeler USA and Accutome.

MISSION:
Analyze ophthalmology and optometry practices using public data, then map clinical workflows to Keeler/Accutome products. Deliver structured, actionable sales intelligence for the field rep.

TOOLS AVAILABLE:
- web_search: Use this to research the practice — find their website, provider list, specialties, locations, job postings, press releases, and any public buying signals. Always search before generating output.

KEELER/ACCUTOME PRODUCT CATALOG:
[PRODUCTS_INJECTED_AT_RUNTIME]

KEY OPHTHALMIC CPT CODES (memorized reference):
92002/92012 — Ophthalmological exam (new/established patient) → any ophthalmic practice
92100 — Serial tonometry → tonometer opportunity
92132/92133/92134 — OCT scanning (anterior/optic nerve/posterior) → imaging equipment
92250 — Fundus photography → retinal camera
76514 — Contact B-scan ultrasound → Accutome B-Scan Pro
76519 — A-scan biometry (IOL calc) → Accutome A-Scan, pre-cataract
92020 — Gonioscopy → slit lamp + gonio lens
92025 — Computerized corneal topography → topographer
92081-92083 — Visual field testing → perimeter
67028 — Intravitreal injection → anti-VEGF workflow, retina
66984 — Cataract extraction → surgical loupes, A-scan
66172/66174 — Filtering glaucoma procedure → glaucoma surgical
92285 — Anterior segment photography → slit lamp camera
68761 — Closure of lacrimal punctum → minor procedure

RULES:
1. Use web_search to research the practice before generating output.
2. NEVER fabricate information. Label inferred data clearly.
3. Only surface ACTIVE ICD-10 and HCPCS codes.
4. Keep output sales-friendly — explain codes in plain English.
5. If data is missing, say so and continue with best-effort logic.
6. Do NOT claim access to ERP, CRM, order history, or internal pricing.

OUTPUT FORMAT — Return ONLY valid JSON with this exact structure. No markdown. No code fences:

{
  "A": {
    "practice_name": "string",
    "location": "string",
    "specialty_focus": "string",
    "publicly_visible_services": ["string"],
    "providers": "string — count and types",
    "buying_signals": ["string — specific signals found"]
  },
  "B": {
    "workflow_summary": "string — 2-3 sentences plain English",
    "steps": [
      {"label": "Disease State", "detail": "string"},
      {"label": "Patient Visit", "detail": "string"},
      {"label": "Diagnostic Testing", "detail": "string"},
      {"label": "Procedure/Treatment", "detail": "string"},
      {"label": "Products Used", "detail": "string"},
      {"label": "Consumables", "detail": "string"},
      {"label": "Replacement Cycle", "detail": "string"}
    ]
  },
  "C": {
    "stages": [
      {"stage": "Pre-Exam", "products": ["string"], "type": "string"},
      {"stage": "Exam", "products": ["string"], "type": "string"},
      {"stage": "Diagnostics", "products": ["string"], "type": "string"},
      {"stage": "Treatment Support", "products": ["string"], "type": "string"},
      {"stage": "Consumables", "products": ["string"], "type": "Daily-Use Consumable|Monthly Reorder|Procedure-Driven"},
      {"stage": "Capital Equipment", "products": ["string"], "type": "Long-Cycle Capital"}
    ]
  },
  "D": {
    "codes": [
      {
        "type": "CPT|HCPCS|ICD-10|NDC",
        "code": "string",
        "description": "string",
        "sales_meaning": "plain English sales relevance — one sentence"
      }
    ]
  },
  "E": {
    "items": [
      {
        "product": "string",
        "class": "Daily-Use Consumable|Monthly Reorder|Procedure-Driven|Annual Maintenance|Long-Cycle Capital|Warranty/Service",
        "notes": "string — reorder frequency or replacement logic"
      }
    ]
  },
  "F": {
    "opportunities": [
      {
        "anchor": "string — the product/workflow already present",
        "cross_sell": "string — the recommended Keeler/Accutome product",
        "rationale": "string — clinical workflow reason"
      }
    ]
  },
  "G": {
    "why_fit": "string — 1-2 sentences on why this account is a good fit",
    "buying_signals": ["string"],
    "top_opportunities": ["string — ranked, most specific first"],
    "confidence": "High|Medium|Low",
    "confidence_reasoning": "string — what data was found vs estimated"
  },
  "H": {
    "action": "string — one clear specific sentence: what the rep should do next",
    "detail": "string — 1-2 sentences of supporting context"
  },
  "meta": {
    "prospect_name": "string",
    "analysis_date": "string — today's date",
    "npi_verified": true,
    "data_quality": "High|Medium|Low",
    "sources_searched": ["string"]
  }
}"""


# ══════════════════════════════════════════════════════════════════════════════
# User Prompt Builder
# ══════════════════════════════════════════════════════════════════════════════

def build_user_prompt(
    practice_name: str,
    npi_number: str,
    location: str,
    website: str,
    specialty: str,
    notes: str,
    npi_data: dict,
    product_catalog: str,
) -> str:

    parts = [
        f"Research and analyze this ophthalmic practice for Keeler/Accutome sales intelligence:\n",
        f"PRACTICE: {practice_name}",
    ]

    if location:
        parts.append(f"LOCATION: {location}")
    if website:
        parts.append(f"WEBSITE: {website}")
    if specialty:
        parts.append(f"SPECIALTY (rep-provided): {specialty}")
    if notes:
        parts.append(f"REP NOTES: {notes}")

    if npi_data and npi_data.get("found"):
        parts.append("\nNPI REGISTRY DATA (confirmed):")
        for rec in npi_data.get("records", [])[:2]:
            parts.append(
                f"  NPI: {rec['npi']} | {rec['org_name']}"
                + (f" (DBA: {rec['dba']})" if rec.get('dba') else "")
            )
            parts.append(f"  Address: {rec['address']}, {rec['city']}, {rec['state']} {rec['zip']}")
            parts.append(f"  Phone: {rec.get('phone','—')}")
            parts.append(f"  Primary Specialty: {rec['primary_specialty']}")
            if rec.get("all_specialties"):
                parts.append(f"  All Specialties: {', '.join(rec['all_specialties'])}")
            parts.append(f"  NPI Type: {rec['type']} | Active since: {rec.get('enumeration_date','?')}")
    else:
        parts.append("\nNPI REGISTRY DATA: Not found — use web search to supplement.")

    icd_context = get_icd10_ophthalmic_context()
    if icd_context:
        parts.append(f"\n{icd_context}")

    hcpcs_context = get_ophthalmic_hcpcs_context()
    if hcpcs_context:
        parts.append(f"\n{hcpcs_context}")

    if CPT_DATA:
        parts.append(f"\nCPT DATA: {len(CPT_DATA):,} codes loaded.")
    else:
        parts.append("\nCPT DATA: Not loaded — use the memorized ophthalmic CPT codes in the system prompt.")

    parts.append(f"\nKEELER/ACCUTOME PRODUCT CATALOG:\n{product_catalog}")

    parts.append(
        "\nINSTRUCTIONS:\n"
        "1. Use web_search to find this practice's website, providers, specialties, services, job postings, and press releases.\n"
        "2. Build the complete A-H analysis based on confirmed public data + best-effort inference.\n"
        "3. Clearly label what is CONFIRMED vs INFERRED.\n"
        "4. Return only valid JSON — no markdown, no code blocks."
    )

    return "\n".join(parts)


# ══════════════════════════════════════════════════════════════════════════════
# Routes
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/")
def index():
    return send_from_directory("static", "index.html")


@app.route("/api/health")
def health():
    username = request.headers.get("X-SCC-User", "unknown")
    return jsonify({
        "status": "ok",
        "user": username,
        "icd10_codes": len(ICD10),
        "hcpcs_codes": len(HCPCS_DATA),
        "cpt_codes": len(CPT_DATA),
        "model": MODEL,
    })


@app.route("/api/analyze", methods=["POST"])
def analyze():
    """SSE streaming analysis. Emits: { type: status|result|error, ... }"""
    form_data = request.get_json() or {}
    username = request.headers.get("X-SCC-User", "dev-user")

    practice_name = form_data.get("practice_name", "").strip()
    npi_number    = form_data.get("npi_number", "").strip()
    location      = form_data.get("location", "").strip()
    website       = form_data.get("website", "").strip()
    specialty     = form_data.get("specialty", "").strip()
    notes         = form_data.get("notes", "").strip()

    if not practice_name:
        return jsonify({"error": "practice_name is required"}), 400

    def generate():
        try:
            # Step 1 — NPI
            yield f"data: {json.dumps({'type':'status','step':1,'message':'Looking up practice in NPI Registry...'})}\n\n"
            npi_data = lookup_npi(
                npi_number=npi_number if npi_number else None,
                org_name=practice_name if not npi_number else None,
                state=location[-2:] if location and len(location) >= 2 else None,
            )

            # Step 2 — Coding context
            yield f"data: {json.dumps({'type':'status','step':2,'message':'Loading clinical coding context (ICD-10 / HCPCS)...'})}\n\n"
            time.sleep(0.1)

            # Step 3 — Product catalog
            yield f"data: {json.dumps({'type':'status','step':3,'message':'Fetching Keeler/Accutome product catalog...'})}\n\n"
            from scraper import get_products, format_for_prompt
            products = get_products()
            product_catalog = format_for_prompt(products)

            # Step 4 — Claude
            yield f"data: {json.dumps({'type':'status','step':4,'message':'Analyzing with Claude AI (web research + clinical mapping)...'})}\n\n"

            system = SYSTEM_PROMPT.replace("[PRODUCTS_INJECTED_AT_RUNTIME]", product_catalog)
            user_prompt = build_user_prompt(
                practice_name=practice_name,
                npi_number=npi_number,
                location=location,
                website=website,
                specialty=specialty,
                notes=notes,
                npi_data=npi_data,
                product_catalog=product_catalog,
            )

            # ── Tool-use loop ──────────────────────────────────────────────
            messages = [{"role": "user", "content": user_prompt}]
            result_text = ""

            for iteration in range(20):
                response = client.messages.create(
                    model=MODEL,
                    max_tokens=16000,
                    tools=[{"type": "web_search_20260209", "name": "web_search"}],
                    system=system,
                    messages=messages,
                )

                block_types = [b.type for b in response.content]
                print(f"[analyze] round={iteration+1} stop={response.stop_reason} blocks={block_types}", flush=True)

                for block in response.content:
                    if block.type == "text":
                        result_text += block.text

                if response.stop_reason in ("end_turn", "max_tokens"):
                    break

                if response.stop_reason == "tool_use":
                    messages.append({"role": "assistant", "content": response.content})
                    tool_results = []
                    for block in response.content:
                        if block.type == "tool_use":
                            tool_results.append({
                                "type": "tool_result",
                                "tool_use_id": block.id,
                                "content": "",
                            })
                    if tool_results:
                        messages.append({"role": "user", "content": tool_results})
                else:
                    print(f"[analyze] unexpected stop_reason: {response.stop_reason}", flush=True)
                    break

            print(f"[analyze] done — result_text={len(result_text)} chars", flush=True)

            if not result_text:
                for i, blk in enumerate(response.content):
                    print(f"[analyze]   block[{i}] type={blk.type}", flush=True)
                raise ValueError("Claude returned no text. Check [analyze] logs above.")

            # Extract JSON — find outermost { } regardless of surrounding prose or fences
            json_start = result_text.find("{")
            json_end   = result_text.rfind("}")
            if json_start == -1 or json_end <= json_start:
                print(f"[analyze] No JSON braces found. First 300 chars: {result_text[:300]!r}", flush=True)
                raise ValueError("Claude response contained no JSON object.")

            candidate = result_text[json_start:json_end + 1]
            print(f"[analyze] JSON candidate: {len(candidate)} chars, starts={candidate[:80]!r}", flush=True)

            parsed = json.loads(candidate)
            parsed.setdefault("meta", {})
            parsed["meta"]["analyzed_by"] = username
            parsed["meta"]["npi_verified"] = bool(npi_data.get("found"))

            yield f"data: {json.dumps({'type':'result','data':parsed})}\n\n"


        except json.JSONDecodeError as e:
            msg = f"Agent returned non-JSON response: {str(e)}"
            print(f"[analyze] JSON error: {msg}", flush=True)
            yield f"data: {json.dumps({'type':'error','message':msg})}\n\n"
        except anthropic.APIError as e:
            msg = f"Claude API error: {str(e)}"
            print(f"[analyze] Anthropic error: {msg}", flush=True)
            yield f"data: {json.dumps({'type':'error','message':msg})}\n\n"
        except Exception as e:
            import traceback
            msg = str(e)
            print(f"[analyze] Exception: {msg}", flush=True)
            traceback.print_exc()
            yield f"data: {json.dumps({'type':'error','message':msg})}\n\n"
        finally:
            yield "data: [DONE]\n\n"

    return Response(
        stream_with_context(generate()),
        content_type="text/event-stream",
        headers={
            "X-Accel-Buffering": "no",
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
        },
    )


# ══════════════════════════════════════════════════════════════════════════════
# Entry point
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    port = os.environ.get("PORT")
    if not port:
        port = 5001
        print("[dev] PORT not set — defaulting to 5001 for local development")
    app.run(host="0.0.0.0", port=int(port), debug=False, threaded=True)
